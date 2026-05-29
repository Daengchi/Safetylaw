"""
Excel 리포트 생성.

시트 구성:
  1. 법규 목록표   — laws.json 전체 법령 현황 (법규명/제정일/개정차수/개정일/시행일/비고)
  2. 신구법비교    — 변경된 법령의 신구조문 비교 (법령명/조문번호/조문명/개정일/시행일자/구법내용/신법내용)

행 색상 (신구법비교):
  신설 조문 (구법내용 없음): 연두
  삭제 조문 (신법내용 없음): 연분홍
  개정 조문 (둘 다 있음):   연노랑
"""
import io
import re
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Color, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_COLOR = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "신설":      "E2EFDA",
    "삭제":      "FCE4D6",
    "개정":      "FFF2CC",
    "law_a":     "FFFFFF",
    "law_b":     "F2F7FF",
    "border":    "BFBFBF",
    # 법규 목록표 계층 색상
    "list_parent": "DAEEF3",   # 상위법 행: 연한 파랑
    "list_sub":    "FFFFFF",   # 하위법 행: 흰색
}

_COMP_HEADERS = ["법령명", "조문번호", "조문명", "개정일", "시행일자", "구법 내용", "신법 내용"]
_COMP_WIDTHS  = [28, 10, 22, 12, 12, 52, 52]

_LIST_HEADERS = ["No.", "법규명", "개정일", "시행일", "개정 조문 수", "비고"]
_LIST_WIDTHS  = [6, 38, 14, 14, 14, 20]

_P_CONTENT_RE = re.compile(r'<P>(.*?)</P>', re.IGNORECASE | re.DOTALL)
_INVALID_SHEET_CHARS = re.compile(r'[/\\?\*\[\]:]')
# XML 1.0에서 허용되지 않는 제어 문자 (법령 API 텍스트에 혼입될 수 있음)
_ILLEGAL_XML_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f\ufffe\uffff]')


def _san(text: str) -> str:
    """XML 비허용 제어 문자 제거."""
    return _ILLEGAL_XML_CHARS.sub('', text) if text else text


def _safe_sheet_name(name: str) -> str:
    return _INVALID_SHEET_CHARS.sub("_", name)[:31]


# ─── 공통 헬퍼 ────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(border_style="thin", color=_COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row: int, col: int, value: str, width: float) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=_COLOR["header_fg"], name="맑은 고딕", size=10)
    cell.fill      = PatternFill("solid", fgColor=_COLOR["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    ws.column_dimensions[get_column_letter(col)].width = width


def _data_cell(ws, row: int, col: int, value, bg: str, wrap: bool = False,
               bold: bool = False) -> None:
    if isinstance(value, str):
        value = _san(value)
    cell = ws.cell(row=row, column=col, value=value)
    if not isinstance(value, CellRichText):
        cell.font = Font(name="맑은 고딕", size=9, bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.border    = _thin_border()


# ─── 신구법비교 시트 (법령별) ─────────────────────────────────────────────────

def _to_rich_text(text: str):
    """<P>...</P> 구간을 빨강 볼드로 변환. P 태그 없으면 str 반환."""
    text = _san(text)
    if not text or '<P>' not in text.upper():
        return text or ""
    blocks = []
    last = 0
    for m in _P_CONTENT_RE.finditer(text):
        if m.start() > last:
            segment = _san(text[last:m.start()])
            if segment:
                blocks.append(TextBlock(InlineFont(rFont="맑은 고딕", sz=9), segment))
        inner = _san(m.group(1))
        if inner:
            blocks.append(TextBlock(
                InlineFont(rFont="맑은 고딕", sz=9, b=True, color=Color(rgb="FFFF0000")),
                inner,
            ))
        last = m.end()
    if last < len(text):
        segment = _san(text[last:])
        if segment:
            blocks.append(TextBlock(InlineFont(rFont="맑은 고딕", sz=9), segment))
    if not blocks:
        return ""
    return CellRichText(*blocks)


def _row_color(art: dict) -> str:
    old = art.get("구법내용", "").strip()
    new = art.get("신법내용", "").strip()
    if not old:
        return _COLOR["신설"]
    if not new:
        return _COLOR["삭제"]
    return _COLOR["개정"]


def _add_group_sheet(ws, group: dict, result_map: dict) -> None:
    """
    그룹 내 모든 법령(법+시행령+시행규칙)을 하나의 시트에 렌더링.
    각 법령은 교대 배경색으로 구분. articles 없는 법령은 안내 메시지 행 삽입.
    """
    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(_COMP_HEADERS, _COMP_WIDTHS), 1):
        _header_cell(ws, 1, col, h, w)

    row       = 2
    color_idx = 0

    for law_info in group.get("laws", []):
        name     = law_info["name"]
        res      = result_map.get(name)
        articles = res.get("articles", []) if res else []

        if not articles:
            # 해당 법령 데이터 없음 → 안내 메시지 한 행 (병합 없이)
            status   = res["status"] if res else "오류"
            date_str = res.get("new_date", "") if res else ""
            if date_str and len(date_str) == 8 and date_str.isdigit():
                date_str = f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
            if res and res.get("타법개정"):
                msg = f"타법개정 (최신 시행일자: {date_str}) — API 신구법비교 미제공"
            elif status == "변경 없음":
                msg = f"변경 없음 (최신 시행일자: {date_str})"
            else:
                msg = "조회 오류 또는 데이터 없음"
            msg_text = f"[{name}] {msg}"
            for col in range(1, len(_COMP_HEADERS) + 1):
                val = msg_text if col == 1 else ""
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(name="맑은 고딕", size=9, italic=True, color="808080")
                c.fill      = PatternFill("solid", fgColor="F2F2F2")
                c.alignment = Alignment(
                    horizontal="center" if col == 1 else "left",
                    vertical="center",
                )
                c.border    = _thin_border()
            ws.row_dimensions[row].height = 22
            row += 1
            continue

        law_bg = _COLOR["law_a"] if color_idx % 2 == 0 else _COLOR["law_b"]
        color_idx += 1

        for art in articles:
            content_bg = _row_color(art)
            old_text   = art.get("구법내용", "")
            new_text   = art.get("신법내용", "")

            _data_cell(ws, row, 1, name,                           law_bg)
            _data_cell(ws, row, 2, art.get("조문번호", ""),        law_bg)
            _data_cell(ws, row, 3, art.get("조문명",   ""),        law_bg)
            _data_cell(ws, row, 4, art.get("개정일",   ""),        law_bg)
            _data_cell(ws, row, 5, art.get("시행일자", ""),        law_bg)
            _data_cell(ws, row, 6, _to_rich_text(old_text), content_bg, wrap=True)
            _data_cell(ws, row, 7, _to_rich_text(new_text), content_bg, wrap=True)

            max_len = max(len(old_text), len(new_text), 1)
            ws.row_dimensions[row].height = max(40, min(max_len // 4, 200))
            row += 1

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:G{max(row - 1, 1)}"


# ─── 법규 목록표 시트 ─────────────────────────────────────────────────────────

def _fmt_date_display(date_str: str) -> str:
    """20251001 → 2025/10/01. 이미 포맷된 문자열은 그대로."""
    if date_str and len(date_str) == 8 and date_str.isdigit():
        return f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
    return date_str or "-"


def _add_law_list_sheet(ws, law_groups: list[dict],
                        article_counts: dict[str, int]) -> None:
    """
    law_groups:     [{"parent": str, "laws": [law_info_dict, ...]}, ...]
    article_counts: {법령명: 개정조문수}  변경 없음=0, 오류=키 없음
    """
    num_cols = len(_LIST_HEADERS)
    last_col = get_column_letter(num_cols)

    # row 1: 제목
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value="안전환경보건 법규 목록표")
    title_cell.font      = Font(name="맑은 고딕", size=15, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # row 2: 빈 행
    ws.row_dimensions[2].height = 12

    # row 3: 최신화 날짜 (E열)
    today = datetime.now().strftime("%Y/%m/%d")
    date_cell = ws.cell(row=3, column=num_cols, value=f"최신화 날짜 : {today}")
    date_cell.font      = Font(name="맑은 고딕", size=9)
    date_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[3].height = 16

    # row 4: 헤더
    ws.row_dimensions[4].height = 28
    for col, (h, w) in enumerate(zip(_LIST_HEADERS, _LIST_WIDTHS), 1):
        _header_cell(ws, 4, col, h, w)

    parent_bg = _COLOR["list_parent"]
    sub_bg    = _COLOR["list_sub"]

    row = 5
    for gi, group in enumerate(law_groups):
        laws        = group.get("laws", [])
        group_start = row

        for i, law in enumerate(laws):
            is_parent = (i == 0)
            bg = parent_bg if is_parent else sub_bg

            개정일 = _fmt_date_display(law.get("공포일자", ""))
            시행일 = _fmt_date_display(law.get("시행일자", ""))
            비고   = law.get("제개정구분명", "")

            cnt     = article_counts.get(law["name"])
            cnt_val = cnt if cnt is not None else "-"

            # No. 열(A=1)은 그룹 처리 후 일괄 적용 — 여기선 빈 셀만 스타일 지정
            c = ws.cell(row=row, column=1, value="")
            c.fill   = PatternFill("solid", fgColor=parent_bg)
            c.border = _thin_border()

            _data_cell(ws, row, 2, law["name"], bg, bold=is_parent)
            _data_cell(ws, row, 3, 개정일,      bg)
            _data_cell(ws, row, 4, 시행일,      bg)
            _data_cell(ws, row, 5, cnt_val,     bg)
            _data_cell(ws, row, 6, 비고,        bg)

            ws.row_dimensions[row].height = 18
            row += 1

        # No. 셀: 그룹 전체 행 병합 후 번호 기입
        group_end = row - 1
        no_cell = ws.cell(row=group_start, column=1, value=gi + 1)
        no_cell.font      = Font(name="맑은 고딕", size=9, bold=True)
        no_cell.fill      = PatternFill("solid", fgColor=parent_bg)
        no_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_cell.border    = _thin_border()

        if group_end > group_start:
            ws.merge_cells(f"A{group_start}:A{group_end}")

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(num_cols)}{max(row - 1, 4)}"


# ─── openpyxl CellRichText 버그 보완 ─────────────────────────────────────────

# 단순 inlineStr 셀 패턴: <is><t...>TEXT</t></is>  (rich text <r> 없음)
_SIMPLE_INLINE_RE = re.compile(
    r'<c(\s[^>]*?)t="inlineStr"([^>]*)><is><t(?:\s[^>]*)?>([^<]*)</t></is></c>',
    re.DOTALL,
)
_WS_FILE_RE = re.compile(r'^xl/worksheets/sheet\d+\.xml$')


def _fix_shared_strings(xlsx_path: str) -> None:
    """
    openpyxl 3.1.x에서 CellRichText 사용 시 sharedStrings.xml이 생성되지 않는 버그 보완.
    단순 텍스트 inlineStr 셀을 sharedStrings 참조로 변환하고 sharedStrings.xml 생성.
    Rich text (<r> 포함) inlineStr 셀은 그대로 유지.
    """
    with open(xlsx_path, 'rb') as f:
        raw = f.read()

    zin = zipfile.ZipFile(io.BytesIO(raw), 'r')
    names = zin.namelist()

    # 이미 sharedStrings.xml이 있으면 처리 불필요
    if 'xl/sharedStrings.xml' in names:
        zin.close()
        return

    # ── 1단계: 변환 대상 문자열 수집 ────────────────────────────────────────
    str_to_idx: dict[str, int] = {}
    str_list:   list[str]      = []

    for name in names:
        if not _WS_FILE_RE.match(name):
            continue
        xml = zin.read(name).decode('utf-8')
        for m in _SIMPLE_INLINE_RE.finditer(xml):
            txt = m.group(3)
            if txt not in str_to_idx:
                str_to_idx[txt] = len(str_list)
                str_list.append(txt)

    if not str_list:
        zin.close()
        return

    # ── 2단계: 워크시트 XML 변환 ─────────────────────────────────────────────
    def _replace(m: re.Match) -> str:
        attrs1, attrs2, txt = m.group(1), m.group(2), m.group(3)
        idx = str_to_idx.get(txt)
        if idx is None:
            return m.group(0)
        return f'<c{attrs1}t="s"{attrs2}><v>{idx}</v></c>'

    zout_buf = io.BytesIO()
    with zipfile.ZipFile(zout_buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)

            if _WS_FILE_RE.match(item.filename):
                data = _SIMPLE_INLINE_RE.sub(_replace, data.decode('utf-8')).encode('utf-8')

            elif item.filename == '[Content_Types].xml':
                s = data.decode('utf-8')
                if 'sharedStrings' not in s:
                    s = s.replace(
                        '</Types>',
                        '<Override PartName="/xl/sharedStrings.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sharedStrings+xml"/></Types>',
                    )
                data = s.encode('utf-8')

            elif item.filename == 'xl/_rels/workbook.xml.rels':
                s = data.decode('utf-8')
                if 'sharedStrings' not in s:
                    rids = re.findall(r'Id="rId(\d+)"', s)
                    nid  = max((int(r) for r in rids), default=0) + 1
                    s = s.replace(
                        '</Relationships>',
                        f'<Relationship Id="rId{nid}" '
                        f'Type="http://schemas.openxmlformats.org/officeDocument/2006'
                        f'/relationships/sharedStrings" Target="sharedStrings.xml"/>'
                        f'</Relationships>',
                    )
                data = s.encode('utf-8')

            zout.writestr(item, data)

        # ── 3단계: sharedStrings.xml 생성 ──────────────────────────────────
        def _esc(s: str) -> str:
            return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        n = len(str_list)
        lines = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' count="{n}" uniqueCount="{n}">',
        ]
        for s in str_list:
            lines.append(f'<si><t xml:space="preserve">{_esc(s)}</t></si>')
        lines.append('</sst>')
        zout.writestr('xl/sharedStrings.xml', '\n'.join(lines).encode('utf-8'))

    zin.close()

    with open(xlsx_path, 'wb') as f:
        f.write(zout_buf.getvalue())


# ─── 진입점 ───────────────────────────────────────────────────────────────────

def generate(results: list[dict], output_path: str,
             law_groups: list[dict] | None = None) -> None:
    """
    results:    _check_law() 반환값 리스트
    law_groups: [{"parent": str, "laws": [law_info, ...]}, ...]  (법규 목록표용)
    """
    wb = Workbook()
    result_map = {r["law_name"]: r for r in results}

    if law_groups:
        article_counts = {r["law_name"]: r.get("article_count", "-") for r in results}

        ws_list = wb.active
        ws_list.title = "1. 법규 목록표"
        _add_law_list_sheet(ws_list, law_groups, article_counts)

        for group in law_groups:
            sheet_ws = wb.create_sheet(_safe_sheet_name(group["parent"]))
            _add_group_sheet(sheet_ws, group, result_map)
    else:
        # law_groups 없는 단독 실행 fallback
        for i, res in enumerate(results):
            name = res["law_name"]
            if i == 0:
                sheet_ws = wb.active
            else:
                sheet_ws = wb.create_sheet()
            sheet_ws.title = _safe_sheet_name(name)
            fake_group = {"parent": name, "laws": [{"name": name}]}
            _add_group_sheet(sheet_ws, fake_group, result_map)

    wb.save(output_path)
    _fix_shared_strings(output_path)
